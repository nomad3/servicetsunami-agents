"""
Media processing utilities for converting raw media bytes into
LLM-compatible message parts (inline_data / text).

Audio transcription (whisper) was moved to apps/code-worker per the
api-image-diet plan (docs/plans/2026-05-18-docker-image-shrink-and-latency.md).
The transcription helpers now live in
``app.services.transcription_client`` and dispatch through Temporal.
"""

import base64
import io
import logging
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ── MIME-type sets ──────────────────────────────────────────────────────────

IMAGE_MIMES = {
    "image/jpeg",
    "image/png",
    "image/webp",
    "image/gif",
    "image/heic",
}

AUDIO_MIMES = {
    "audio/ogg",
    "audio/mpeg",
    "audio/mp4",
    "audio/wav",
    "audio/webm",
    "audio/aac",
}

PDF_MIMES = {"application/pdf"}

SPREADSHEET_MIMES = {
    "text/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

# ── Size limits (bytes) ────────────────────────────────────────────────────

MAX_IMAGE_SIZE = 10 * 1024 * 1024   # 10 MB
MAX_AUDIO_SIZE = 25 * 1024 * 1024   # 25 MB
MAX_PDF_SIZE = 20 * 1024 * 1024     # 20 MB
MAX_SPREADSHEET_SIZE = 10 * 1024 * 1024  # 10 MB

# ── Default prompts ────────────────────────────────────────────────────────

DEFAULT_IMAGE_PROMPT = "The user sent this image. Describe what you see and respond helpfully."
DEFAULT_AUDIO_PROMPT = "The user sent this voice message. Transcribe and respond to it."
DEFAULT_PDF_PROMPT = "The user sent a PDF. Please review and respond."

# ── PDF extraction limits ──────────────────────────────────────────────────

MAX_PDF_PAGES = 30
MAX_PDF_CHARS = 50_000


# ── Public API ──────────────────────────────────────────────────────────────

def classify_media(mime_type: str) -> str:
    """Classify a MIME type into image, audio, pdf, or unsupported."""
    # Clean the mime_type (handles "audio/ogg; codecs=opus")
    clean = mime_type.split(";")[0].strip().lower()

    if clean in IMAGE_MIMES:
        return "image"
    if clean in AUDIO_MIMES or clean.startswith("audio/"):
        return "audio"
    if clean in PDF_MIMES:
        return "pdf"
    if clean in SPREADSHEET_MIMES:
        return "spreadsheet"
    return "unsupported"


def build_media_parts(
    media_bytes: bytes,
    mime_type: str,
    caption: str = "",
    filename: str = "",
    *,
    precomputed_transcript: Optional[str] = None,
) -> Tuple[List[Dict], Dict]:
    """
    Convert raw media bytes into LLM-compatible message parts.

    For audio: if ``precomputed_transcript`` is supplied the function will
    NOT call ``transcribe_bytes_sync`` (which blocks the event loop and is
    only safe from sync handlers). Async callers MUST resolve the
    transcript via ``await transcription_client.transcribe_async(...)``
    first and pass it through.

    Returns:
        (parts, attachment_meta)
        - parts: list of message-part dicts
        - attachment_meta: metadata dict with type, mime_type, size_bytes, filename
    """
    clean_mime = mime_type.split(";")[0].strip().lower()
    media_class = classify_media(clean_mime)

    if media_class == "unsupported":
        raise ValueError(f"Unsupported media type: {mime_type}")

    size = len(media_bytes)

    # Enforce size limits
    if media_class == "image" and size > MAX_IMAGE_SIZE:
        raise ValueError(
            f"Image too large: {size} bytes (max {MAX_IMAGE_SIZE} bytes)"
        )
    if media_class == "audio" and size > MAX_AUDIO_SIZE:
        raise ValueError(
            f"Audio too large: {size} bytes (max {MAX_AUDIO_SIZE} bytes)"
        )
    if media_class == "pdf" and size > MAX_PDF_SIZE:
        raise ValueError(
            f"PDF too large: {size} bytes (max {MAX_PDF_SIZE} bytes)"
        )
    if media_class == "spreadsheet" and size > MAX_SPREADSHEET_SIZE:
        raise ValueError(
            f"Spreadsheet too large: {size} bytes (max {MAX_SPREADSHEET_SIZE} bytes)"
        )

    # Build parts by media class
    if media_class == "image":
        parts = _build_image_parts(media_bytes, clean_mime, caption)
    elif media_class == "audio":
        parts = _build_audio_parts(
            media_bytes,
            clean_mime,
            caption,
            precomputed_transcript=precomputed_transcript,
        )
    elif media_class == "pdf":
        parts = _build_pdf_parts(media_bytes, caption, filename)
    elif media_class == "spreadsheet":
        parts = _build_spreadsheet_parts(media_bytes, clean_mime, caption, filename)
    else:
        raise ValueError(f"Unsupported media type: {mime_type}")

    attachment_meta = {
        "type": media_class,
        "mime_type": clean_mime,
        "size_bytes": size,
        "filename": filename,
    }

    return parts, attachment_meta


# ── Internal helpers ────────────────────────────────────────────────────────

def _build_image_parts(
    image_bytes: bytes,
    mime_type: str,
    caption: str,
) -> List[Dict]:
    """Base64-encode an image and return inline_data + text parts."""
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    text = caption if caption else DEFAULT_IMAGE_PROMPT

    return [
        {"inline_data": {"mime_type": mime_type, "data": b64}},
        {"text": text},
    ]


def _build_audio_parts(
    audio_bytes: bytes,
    mime_type: str,
    caption: str,
    *,
    precomputed_transcript: Optional[str] = None,
) -> List[Dict]:
    """Transcribe audio via the code-worker transcription workflow; fall back to inline_data.

    When ``precomputed_transcript`` is provided we skip the sync workflow
    dispatch entirely — the async caller already resolved the transcript
    via ``transcribe_async``. This avoids blocking the event loop with
    ``transcribe_bytes_sync``'s ThreadPoolExecutor bridge.
    """
    transcript: Optional[str] = precomputed_transcript
    if transcript is None:
        try:
            from app.services.transcription_client import transcribe_bytes_sync

            transcript = transcribe_bytes_sync(audio_bytes)
        except Exception:
            logger.exception("Inline audio transcription dispatch failed; falling back to inline_data")

    if transcript:
        prompt = f"[Voice message transcription]: {transcript}"
        if caption:
            prompt += f"\n[Caption: {caption}]"
        return [{"text": prompt}]

    # Fallback: send as inline data for the LLM to handle
    b64 = base64.b64encode(audio_bytes).decode("utf-8")
    text = caption if caption else DEFAULT_AUDIO_PROMPT
    return [
        {"inline_data": {"mime_type": mime_type, "data": b64}},
        {"text": text},
    ]


def _build_pdf_parts(
    pdf_bytes: bytes,
    caption: str,
    filename: str,
) -> List[Dict]:
    """Extract text from a PDF with pdfplumber and return a text part."""
    import pdfplumber

    extracted_pages: List[str] = []

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            pages_to_read = pdf.pages[:MAX_PDF_PAGES]
            for i, page in enumerate(pages_to_read):
                page_text = page.extract_text()
                if page_text:
                    extracted_pages.append(f"--- Page {i + 1} ---\n{page_text}")

            total_pages = len(pdf.pages)
    except Exception:
        logger.exception("Failed to extract text from PDF")
        extracted_pages = ["[Could not extract PDF text]"]
        total_pages = 0

    full_text = "\n\n".join(extracted_pages)

    # Truncate if too long
    truncated = False
    if len(full_text) > MAX_PDF_CHARS:
        full_text = full_text[:MAX_PDF_CHARS]
        truncated = True

    # Build header
    header_parts = []
    if filename:
        header_parts.append(f"Filename: {filename}")
    header_parts.append(f"Pages: {min(total_pages, MAX_PDF_PAGES)}/{total_pages}")
    if truncated:
        header_parts.append(f"(truncated to {MAX_PDF_CHARS} chars)")
    header = " | ".join(header_parts)

    prompt = caption if caption else DEFAULT_PDF_PROMPT

    content = f"{prompt}\n\n--- PDF Content ({header}) ---\n{full_text}"

    return [{"text": content}]


MAX_SPREADSHEET_CHARS = 50_000


def _build_spreadsheet_parts(
    file_bytes: bytes,
    mime_type: str,
    caption: str,
    filename: str,
) -> List[Dict]:
    """Extract text from CSV or Excel and return a text part."""
    try:
        if mime_type == "text/csv":
            text_content = file_bytes.decode("utf-8", errors="replace")
        else:
            # Excel file — extract with openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(c) if c is not None else "" for c in row]
                    if any(v for v in row_vals):
                        rows.append(",".join(row_vals))
                if rows:
                    sheets.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
            text_content = "\n\n".join(sheets)

        truncated = False
        if len(text_content) > MAX_SPREADSHEET_CHARS:
            text_content = text_content[:MAX_SPREADSHEET_CHARS]
            truncated = True

        header_parts = []
        if filename:
            header_parts.append(f"Filename: {filename}")
        if truncated:
            header_parts.append(f"(truncated to {MAX_SPREADSHEET_CHARS} chars)")
        header = " | ".join(header_parts) if header_parts else "Spreadsheet"

        prompt = caption if caption else "The user sent a spreadsheet. Please review and respond."
        content = f"{prompt}\n\n--- Spreadsheet Content ({header}) ---\n{text_content}"

        return [{"text": content}]
    except Exception:
        logger.exception("Failed to extract spreadsheet content")
        return [{"text": f"{caption or 'The user sent a spreadsheet.'}\n\n[Could not extract spreadsheet content from {filename}]"}]


# ── Back-compat shims ──────────────────────────────────────────────────────
#
# The pre-migration module exposed ``transcribe_audio_bytes`` and
# ``transcribe_audio_path``. A handful of callers (whatsapp_service.py,
# robot.py, the chat file-upload endpoint) imported these directly.
# Keep the names alive but route through the transcription_client so old
# call sites don't break — new code should import from
# ``app.services.transcription_client`` directly.


def transcribe_audio_bytes(audio_bytes: bytes) -> Optional[str]:
    """Compatibility shim — dispatches through the code-worker workflow."""
    from app.services.transcription_client import transcribe_bytes_sync

    return transcribe_bytes_sync(audio_bytes)


def transcribe_audio_path(path: str) -> Optional[str]:
    """Compatibility shim — reads the file and dispatches through the workflow.

    Kept for callers that already wrote a temp file. Newer code paths should
    hand the bytes straight to ``transcribe_bytes_sync`` to avoid the
    extra read.
    """
    from app.services.transcription_client import transcribe_bytes_sync

    try:
        with open(path, "rb") as fh:
            data = fh.read()
    except OSError:
        logger.exception("transcribe_audio_path: failed to read %s", path)
        return None
    return transcribe_bytes_sync(data)
