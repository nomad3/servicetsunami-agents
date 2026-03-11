"""Report generation endpoints — Excel practice-performance reports.

Two entry points:
- POST /generate  — JWT-authenticated (web UI)
- POST /internal/generate — header-authenticated (ADK → API callbacks)
- GET  /download/{file_id} — serves the generated .xlsx file
"""

import logging
import time
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import List

from fastapi import APIRouter, Depends, Header, HTTPException, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.api import deps
from app.models.user import User

logger = logging.getLogger(__name__)
router = APIRouter()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
REPORTS_DIR = Path("/tmp/servicetsunami_reports")
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
REPORT_TTL_HOURS = 24

# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------


class ProviderData(BaseModel):
    name: str
    production: float = 0.0
    collections: float = 0.0
    visits: int = 0
    presented: float = 0.0
    accepted: float = 0.0


class HygieneData(BaseModel):
    name: str
    production: float = 0.0
    collections: float = 0.0
    visits: int = 0
    capacity: int = 0
    utilized: int = 0
    reappointment_rate: float = 0.0


class ProductionData(BaseModel):
    period: str = ""
    practice_name: str = "Practice"
    doctors: List[ProviderData] = Field(default_factory=list)
    specialists: List[ProviderData] = Field(default_factory=list)
    hygienists: List[HygieneData] = Field(default_factory=list)
    adjustments: float = 0.0
    write_offs: float = 0.0
    refunds: float = 0.0


class ReportRequest(BaseModel):
    report_type: str = "monthly_performance"
    period: str = ""
    data: ProductionData = Field(default_factory=ProductionData)


class ReportResponse(BaseModel):
    file_id: str
    download_url: str
    expires_at: str


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

FMT_CURRENCY = '$#,##0.00'
FMT_PERCENT = '0.0%'
FMT_NUMBER = '#,##0'


def _apply_header_row(ws, row: int, values: list, col_start: int = 1):
    """Write a styled header row."""
    for idx, val in enumerate(values, start=col_start):
        cell = ws.cell(row=row, column=idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_subheader_row(ws, row: int, values: list, col_start: int = 1):
    """Write a styled sub-header row."""
    for idx, val in enumerate(values, start=col_start):
        cell = ws.cell(row=row, column=idx, value=val)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_currency(ws, row: int, col: int, value: float):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = FMT_CURRENCY
    return cell


def _write_percent(ws, row: int, col: int, value: float):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = FMT_PERCENT
    return cell


def _write_number(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = FMT_NUMBER
    return cell


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _build_excel(data: ProductionData, period: str) -> str:
    """Build a formatted Excel workbook and return the file path (without tenant prefix)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Practice Performance"

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col_letter in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 18

    row = 1

    # ── Title ──────────────────────────────────────────────────────────────
    title_cell = ws.cell(row=row, column=1, value=f"{data.practice_name} — Performance Report")
    title_cell.font = TITLE_FONT
    row += 1
    if period:
        ws.cell(row=row, column=1, value=f"Period: {period}")
    row += 2

    # ── Section 1: Production & Collections ────────────────────────────────
    ws.cell(row=row, column=1, value="PRODUCTION & COLLECTIONS").font = Font(
        name="Calibri", size=12, bold=True,
    )
    row += 1
    _apply_header_row(ws, row, ["Provider", "Gross Production", "Net Production", "Collections", "% Net Production"])
    row += 1

    all_providers: list[ProviderData] = []

    # Doctors
    if data.doctors:
        _apply_subheader_row(ws, row, ["Doctors", "", "", "", ""])
        row += 1
        for doc in data.doctors:
            ws.cell(row=row, column=1, value=doc.name)
            _write_currency(ws, row, 2, doc.production)
            net = doc.production - (data.adjustments + data.write_offs) / max(len(data.doctors) + len(data.specialists), 1)
            _write_currency(ws, row, 3, net)
            _write_currency(ws, row, 4, doc.collections)
            pct = (net / doc.production) if doc.production else 0.0
            _write_percent(ws, row, 5, pct)
            all_providers.append(doc)
            row += 1

    # Specialists
    if data.specialists:
        _apply_subheader_row(ws, row, ["Specialists", "", "", "", ""])
        row += 1
        for sp in data.specialists:
            ws.cell(row=row, column=1, value=sp.name)
            _write_currency(ws, row, 2, sp.production)
            net = sp.production - (data.adjustments + data.write_offs) / max(len(data.doctors) + len(data.specialists), 1)
            _write_currency(ws, row, 3, net)
            _write_currency(ws, row, 4, sp.collections)
            pct = (net / sp.production) if sp.production else 0.0
            _write_percent(ws, row, 5, pct)
            all_providers.append(sp)
            row += 1

    # Hygiene
    if data.hygienists:
        _apply_subheader_row(ws, row, ["Hygiene", "", "", "", ""])
        row += 1
        for hyg in data.hygienists:
            ws.cell(row=row, column=1, value=hyg.name)
            _write_currency(ws, row, 2, hyg.production)
            _write_currency(ws, row, 3, hyg.production)  # hygiene = gross
            _write_currency(ws, row, 4, hyg.collections)
            pct = (hyg.production / hyg.production) if hyg.production else 0.0
            _write_percent(ws, row, 5, pct)
            row += 1

    # Totals
    total_gross = sum(p.production for p in all_providers) + sum(h.production for h in data.hygienists)
    total_net = total_gross - data.adjustments - data.write_offs - data.refunds
    total_collections = sum(p.collections for p in all_providers) + sum(h.collections for h in data.hygienists)
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = SUBHEADER_FONT
    _write_currency(ws, row, 2, total_gross)
    _write_currency(ws, row, 3, total_net)
    _write_currency(ws, row, 4, total_collections)
    _write_percent(ws, row, 5, (total_net / total_gross) if total_gross else 0.0)
    row += 2

    # ── Section 2: Patient Visits ──────────────────────────────────────────
    ws.cell(row=row, column=1, value="PATIENT VISITS").font = Font(
        name="Calibri", size=12, bold=True,
    )
    row += 1
    _apply_header_row(ws, row, ["Provider", "Visits"])
    row += 1

    if data.doctors:
        _apply_subheader_row(ws, row, ["Doctors", ""])
        row += 1
        for doc in data.doctors:
            ws.cell(row=row, column=1, value=doc.name)
            _write_number(ws, row, 2, doc.visits)
            row += 1

    if data.specialists:
        _apply_subheader_row(ws, row, ["Specialists", ""])
        row += 1
        for sp in data.specialists:
            ws.cell(row=row, column=1, value=sp.name)
            _write_number(ws, row, 2, sp.visits)
            row += 1

    if data.hygienists:
        _apply_subheader_row(ws, row, ["Hygienists", ""])
        row += 1
        for hyg in data.hygienists:
            ws.cell(row=row, column=1, value=hyg.name)
            _write_number(ws, row, 2, hyg.visits)
            row += 1

    total_visits = (
        sum(d.visits for d in data.doctors)
        + sum(s.visits for s in data.specialists)
        + sum(h.visits for h in data.hygienists)
    )
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = SUBHEADER_FONT
    _write_number(ws, row, 2, total_visits)
    row += 2

    # ── Section 3: Gross Production By Provider ────────────────────────────
    ws.cell(row=row, column=1, value="GROSS PRODUCTION BY PROVIDER").font = Font(
        name="Calibri", size=12, bold=True,
    )
    row += 1
    _apply_header_row(ws, row, ["Provider", "Gross Production", "% of Total"])
    row += 1
    everyone = [(p.name, p.production) for p in data.doctors + data.specialists] + [
        (h.name, h.production) for h in data.hygienists
    ]
    for name, prod in everyone:
        ws.cell(row=row, column=1, value=name)
        _write_currency(ws, row, 2, prod)
        _write_percent(ws, row, 3, (prod / total_gross) if total_gross else 0.0)
        row += 1
    row += 1

    # ── Section 4: Production Per Visit ────────────────────────────────────
    ws.cell(row=row, column=1, value="PRODUCTION PER VISIT").font = Font(
        name="Calibri", size=12, bold=True,
    )
    row += 1
    _apply_header_row(ws, row, ["Provider", "Production", "Visits", "Per Visit"])
    row += 1
    for name, prod in everyone:
        visits = 0
        for p in data.doctors + data.specialists:
            if p.name == name:
                visits = p.visits
                break
        else:
            for h in data.hygienists:
                if h.name == name:
                    visits = h.visits
                    break
        ws.cell(row=row, column=1, value=name)
        _write_currency(ws, row, 2, prod)
        _write_number(ws, row, 3, visits)
        _write_currency(ws, row, 4, (prod / visits) if visits else 0.0)
        row += 1
    row += 1

    # ── Section 5: Case Acceptance ─────────────────────────────────────────
    ws.cell(row=row, column=1, value="CASE ACCEPTANCE").font = Font(
        name="Calibri", size=12, bold=True,
    )
    row += 1
    _apply_header_row(ws, row, ["Provider", "Presented", "Accepted", "Acceptance Rate"])
    row += 1
    for prov in data.doctors + data.specialists:
        ws.cell(row=row, column=1, value=prov.name)
        _write_currency(ws, row, 2, prov.presented)
        _write_currency(ws, row, 3, prov.accepted)
        rate = (prov.accepted / prov.presented) if prov.presented else 0.0
        _write_percent(ws, row, 4, rate)
        row += 1
    total_presented = sum(p.presented for p in data.doctors + data.specialists)
    total_accepted = sum(p.accepted for p in data.doctors + data.specialists)
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = SUBHEADER_FONT
    _write_currency(ws, row, 2, total_presented)
    _write_currency(ws, row, 3, total_accepted)
    _write_percent(ws, row, 4, (total_accepted / total_presented) if total_presented else 0.0)
    row += 2

    # ── Section 6: Recare ──────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="RECARE").font = Font(
        name="Calibri", size=12, bold=True,
    )
    row += 1
    _apply_header_row(ws, row, ["Hygienist", "Capacity", "Utilized", "Utilization %", "Reappointment Rate"])
    row += 1
    for hyg in data.hygienists:
        ws.cell(row=row, column=1, value=hyg.name)
        _write_number(ws, row, 2, hyg.capacity)
        _write_number(ws, row, 3, hyg.utilized)
        _write_percent(ws, row, 4, (hyg.utilized / hyg.capacity) if hyg.capacity else 0.0)
        _write_percent(ws, row, 5, hyg.reappointment_rate)
        row += 1

    # Save to temp file — file_id only (tenant prefix added by caller)
    file_id = str(uuid.uuid4())
    return wb, file_id


# ---------------------------------------------------------------------------
# Cleanup
# ---------------------------------------------------------------------------

def _cleanup_expired_reports():
    """Remove report files older than REPORT_TTL_HOURS."""
    if not REPORTS_DIR.exists():
        return
    cutoff = time.time() - (REPORT_TTL_HOURS * 3600)
    for fpath in REPORTS_DIR.iterdir():
        if fpath.is_file() and fpath.stat().st_mtime < cutoff:
            try:
                fpath.unlink()
                logger.info("Removed expired report: %s", fpath.name)
            except OSError:
                pass


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/generate", response_model=ReportResponse)
def generate_report(
    payload: ReportRequest,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """Generate an Excel practice-performance report (JWT-authenticated)."""
    tenant_id = str(current_user.tenant_id)
    return _do_generate(payload, tenant_id)


@router.post("/internal/generate", response_model=ReportResponse)
def generate_report_internal(
    payload: ReportRequest,
    x_tenant_id: str = Header(...),
):
    """Generate an Excel report — internal endpoint for ADK callbacks (no JWT)."""
    return _do_generate(payload, x_tenant_id)


@router.get("/download/{file_id}")
def download_report(
    file_id: str,
    tenant_id: str = Query(...),
):
    """Serve a previously generated Excel report for download."""
    # Sanitise inputs
    if ".." in file_id or "/" in file_id or "\\" in file_id:
        raise HTTPException(status_code=400, detail="Invalid file_id")
    if ".." in tenant_id or "/" in tenant_id or "\\" in tenant_id:
        raise HTTPException(status_code=400, detail="Invalid tenant_id")

    file_name = f"{tenant_id}_{file_id}.xlsx"
    file_path = REPORTS_DIR / file_name

    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Report not found or expired")

    return FileResponse(
        path=str(file_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"report_{file_id}.xlsx",
    )


# ---------------------------------------------------------------------------
# Shared logic
# ---------------------------------------------------------------------------

def _do_generate(payload: ReportRequest, tenant_id: str) -> dict:
    """Build the Excel file, persist it, return metadata."""
    _cleanup_expired_reports()

    wb, file_id = _build_excel(payload.data, payload.period)

    file_name = f"{tenant_id}_{file_id}.xlsx"
    file_path = REPORTS_DIR / file_name
    wb.save(str(file_path))
    logger.info("Generated report %s for tenant %s", file_id, tenant_id)

    expires_at = (datetime.utcnow() + timedelta(hours=REPORT_TTL_HOURS)).isoformat()
    download_url = f"/api/v1/reports/download/{file_id}?tenant_id={tenant_id}"

    return {
        "file_id": file_id,
        "download_url": download_url,
        "expires_at": expires_at,
    }
